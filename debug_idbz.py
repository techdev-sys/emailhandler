import openpyxl
try:
    wb = openpyxl.load_workbook('IDBZ.xlsx', read_only=True)
    print(f"Sheets: {wb.sheetnames}")
    for name in wb.sheetnames[:2]:
        s = wb[name]
        print(f"Sheet '{name}' B1: {s['B1'].value}")
except Exception as e:
    print(f"Error: {e}")
