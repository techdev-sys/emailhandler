import zipfile
import re
import datetime
import openpyxl

def find_date(filepath):
    print(f"Scanning {filepath} for date...")
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active
        
        # Scan first 20 rows, first 10 cols
        for r in range(1, 21):
            for c in range(1, 11):
                val = sheet.cell(row=r, column=c).value
                if val:
                    s_val = str(val).upper()
                    if "DATE" in s_val:
                        print(f"Found 'DATE' at {r},{c}: {val}")
                        # Check neighbors for actual date
                        # right
                        val_r = sheet.cell(row=r, column=c+1).value
                        print(f"  -> Right ({r},{c+1}): {val_r} (Type: {type(val_r)})")
                        
                        # below
                        val_d = sheet.cell(row=r+1, column=c).value
                        print(f"  -> Below ({r+1},{c}): {val_d} (Type: {type(val_d)})")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    find_date('FIRSTCAPITAL.xlsx')
    find_date('IDBZ.xlsx')
