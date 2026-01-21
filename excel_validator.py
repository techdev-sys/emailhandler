import os
import zipfile
import re
import datetime

# Native Excel Optimizers
import openpyxl
import xlrd
try:
    from pyxlsb import open_workbook as open_xlsb
    HAS_PYXLSB = True
except ImportError:
    HAS_PYXLSB = False

# --- REGULATORY CLASSIFICATION RULES ---
DETECTION_RULES = {
    "BSD2_3": ["assets", "liabilities", "capital", "equity", "income", "expenses", "profit", "loss", "balance sheet", "loans and advances", "deposits"],
    "BSD4": ["currency", "foreign", "exchange", "usd", "zar", "gbp", "eur", "net open position", "foreign assets", "foreign liabilities"]
}

def analyze_return(filepath, user_selected_mode=None):
    """
    DETERMINISTIC REGULATORY CLASSIFICATION.
    Priority: Strict BSD2_3 Composite Rule (Final Authoritative) > BSD4 Keyword Rule.
    """
    result = {
        'type': 'UNKNOWN',
        'status': 'REJECTED',
        'reason': 'No structural match found.',
        'sheets': [],
        'scores': {'BSD2_3': 0, 'BSD4': 0},
        'matches': {'BSD2_3': [], 'BSD4': []},
        'date': None,
        'bank_name': None,
        'conflict': None
    }

    try:
        if not os.path.exists(filepath):
            result['reason'] = "File not found."
            return result
        
        ext = filepath.lower()
        if not ext.endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')):
            result['reason'] = "Unsupported file format."
            return result

        # --- 1. COMPOSITE BSD2_3 RULE (AUTHORITATIVE) ---
        bsd23_candidate = False
        sheet_details = []
        
        # We need to peek at sheet names and B1 values
        if ext.endswith(('.xlsx', '.xlsm')):
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            if len(sheet_names) >= 2:
                # Check Sheet 1 (BSD2)
                s1_name = sheet_names[0].upper()
                s1 = wb[sheet_names[0]]
                b1_val_1 = str(s1['B1'].value).strip().upper() if s1['B1'].value else ""
                
                # Check Sheet 2 (BSD3)
                s2_name = sheet_names[1].upper()
                s2 = wb[sheet_names[1]]
                b1_val_2 = str(s2['B1'].value).strip().upper() if s2['B1'].value else ""

                if "FORM BSD2" in s1_name and b1_val_1 == "BSD2":
                    sheet_details.append("Form BSD2 -> VALID")
                else: sheet_details.append("Form BSD2 -> INVALID OR MISSING")

                if "FORM BSD3" in s2_name and b1_val_2 == "BSD3":
                    sheet_details.append("Form BSD3 -> VALID")
                else: sheet_details.append("Form BSD3 -> INVALID OR MISSING")

                if len(sheet_details) == 2 and all("VALID" in d for d in sheet_details):
                    bsd23_candidate = True
            wb.close()
        
        elif ext.endswith('.xls'):
            wb = xlrd.open_workbook(filepath)
            if wb.nsheets >= 2:
                s1 = wb.sheet_by_index(0)
                s1_name = s1.name.upper()
                b1_val_1 = str(s1.cell_value(0, 1)).strip().upper() if s1.nrows > 0 and s1.ncols > 1 else ""
                
                s2 = wb.sheet_by_index(1)
                s2_name = s2.name.upper()
                b1_val_2 = str(s2.cell_value(0, 1)).strip().upper() if s2.nrows > 0 and s2.ncols > 1 else ""

                if "FORM BSD2" in s1_name and b1_val_1 == "BSD2":
                    sheet_details.append("Form BSD2 -> VALID")
                else: sheet_details.append("Form BSD2 -> INVALID OR MISSING")

                if "FORM BSD3" in s2_name and b1_val_2 == "BSD3":
                    sheet_details.append("Form BSD3 -> VALID")
                else: sheet_details.append("Form BSD3 -> INVALID OR MISSING")

                if len(sheet_details) == 2 and all("VALID" in d for d in sheet_details):
                    bsd23_candidate = True
        
        elif ext.endswith('.xlsb') and HAS_PYXLSB:
            with open_xlsb(filepath) as wb:
                if len(wb.sheets) >= 2:
                    # Sheet 1
                    s1_name = wb.sheets[0].upper()
                    with wb.get_sheet(s1_name) as s1:
                        b1_val_1 = ""
                        for row in s1.rows():
                            b1_val_1 = str(row[1].v).strip().upper() if len(row) > 1 and row[1].v else ""
                            break # Only row 1
                    # Sheet 2
                    s2_name = wb.sheets[1].upper()
                    with wb.get_sheet(s2_name) as s2:
                        b1_val_2 = ""
                        for row in s2.rows():
                            b1_val_2 = str(row[1].v).strip().upper() if len(row) > 1 and row[1].v else ""
                            break # Only row 1
                    
                    if "FORM BSD2" in s1_name and b1_val_1 == "BSD2":
                        sheet_details.append("Form BSD2 -> VALID")
                    else: sheet_details.append("Form BSD2 -> INVALID OR MISSING")

                    if "FORM BSD3" in s2_name and b1_val_2 == "BSD3":
                        sheet_details.append("Form BSD3 -> VALID")
                    else: sheet_details.append("Form BSD3 -> INVALID OR MISSING")

                    if len(sheet_details) == 2 and all("VALID" in d for d in sheet_details):
                        bsd23_candidate = True

        if bsd23_candidate:
            result['type'] = 'BSD2_3'
            result['status'] = 'ACCEPTED'
            result['reason'] = 'Validated COMPOSITE BSD2_3 structure.'
            result['sheets'] = sheet_details
            
            # Metadata scan (get rows from first sheet)
            _, raw_rows_1 = get_sheet_content(filepath, sheet_idx=0, max_rows=20)
            result['date'] = extract_date_from_rows(raw_rows_1)
            result['bank_name'] = extract_bank_name(filepath)
            return result

        # --- 2. FALLBACK TO BSD4 RULES (KEYWORD SCORES) ---
        # Read first sheet, first 20 rows
        content_text, raw_rows = get_sheet_content(filepath, sheet_idx=0, max_rows=20)
        content_block = content_text.lower()
        
        for rtype, keywords in DETECTION_RULES.items():
            for kw in keywords:
                if kw in content_block:
                    result['scores'][rtype] += 1
                    result['matches'][rtype].append(kw)

        score23 = result['scores']['BSD2_3']
        score4 = result['scores']['BSD4']

        detected_type = 'UNKNOWN'
        if score4 > score23:
            detected_type = 'BSD4'
        elif score23 >= 2 and score4 == 0:
            detected_type = 'UNKNOWN' # Must fail if not composite

        result['type'] = detected_type
        
        if detected_type == 'BSD4':
            result['status'] = 'ACCEPTED'
            result['reason'] = 'Validated BSD4 via keyword dominance.'
        else:
            result['status'] = 'REJECTED'
            result['reason'] = f"Invalid BSD2_3 submission: Workbook must contain Form BSD2 (B1=BSD2) and Form BSD3 (B1=BSD3) in the same file." if score23 > score4 else "Unknown return structure."

        # Metadata
        result['date'] = extract_date_from_rows(raw_rows)
        result['bank_name'] = extract_bank_name(filepath, content_block)

        # 4. VERIFICATION RULE (Conflict check)
        if user_selected_mode and user_selected_mode != 'ALL':
            if detected_type != 'UNKNOWN' and detected_type != user_selected_mode:
                result['conflict'] = f"Selected {user_selected_mode} but file structure matches {detected_type}."

    except Exception as e:
        result['reason'] = f"Engine Error: {str(e)}"
    
    return result

def get_sheet_content(filepath, sheet_idx=0, max_rows=20):
    content = ""
    rows_data = []
    ext = filepath.lower()
    try:
        if ext.endswith(('.xlsx', '.xlsm')):
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            if sheet_idx < len(wb.sheetnames):
                sheet = wb[wb.sheetnames[sheet_idx]]
                for i, row in enumerate(sheet.iter_rows(max_row=max_rows, max_col=15)):
                    rv = [str(c.value).strip() if c.value is not None else "" for c in row]
                    rows_data.append(rv)
                    content += " ".join(rv) + " "
            wb.close()
        elif ext.endswith('.xls'):
            wb = xlrd.open_workbook(filepath)
            if sheet_idx < wb.nsheets:
                sheet = wb.sheet_by_index(sheet_idx)
                for r in range(min(max_rows, sheet.nrows)):
                    rv = [str(sheet.cell_value(r, c)).strip() for c in range(min(15, sheet.ncols))]
                    rows_data.append(rv)
                    content += " ".join(rv) + " "
        elif ext.endswith('.xlsb') and HAS_PYXLSB:
            with open_xlsb(filepath) as wb:
                if sheet_idx < len(wb.sheets):
                    with wb.get_sheet(wb.sheets[sheet_idx]) as sheet:
                        for i, row in enumerate(sheet.rows()):
                            if i >= max_rows: break
                            rv = [str(c.v).strip() if c.v is not None else "" for c in list(row)[:15]]
                            rows_data.append(rv)
                            content += " ".join(rv) + " "
    except: pass
    return content, rows_data


def extract_bank_name(filepath, content_block=None):
    KNOWN_BANKS = [
        "EMPOWERBANK", "BANCABC", "STANBIC", "FBCCROWN", "AFC", "SUCCESS", 
        "TIMEBANK", "GETBUCKS", "STEWARD", "ACL", "NMB", "NBS", "ZWMB",
        "FIRSTCAPITAL", "METBANK", "MUKURU", "INNBUCKS", "CBZ",
        "NEDBANK", "ECOBANK", "IDBZ", "CABS", "ZBBANK", "ZBBS", "POSB", "FBCBS",
        "FBC BANK", "FBC BUILDING", "ZB BANK", "ZB BUILDING", "AGRIBANK", "AGRICULTURAL"
    ]
    
    BANK_MAPPINGS = {
        "FBC BANK": "FBCBANK", "FBC BUILDING": "FBCBS", 
        "ZB BANK": "ZBBANK", "ZB BUILDING": "ZBBS",
        "AGRIBANK": "AFC", "AGRICULTURAL": "AFC"
    }

    if content_block:
        for bank in KNOWN_BANKS:
            if bank in content_block:
                return BANK_MAPPINGS.get(bank, bank)
                
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            if 'xl/sharedStrings.xml' in z.namelist():
                xml = z.read('xl/sharedStrings.xml').decode('utf-8', errors='ignore').upper()
                for bank in KNOWN_BANKS:
                    if bank in xml: return BANK_MAPPINGS.get(bank, bank)
    except: pass
    return None

def extract_date_from_rows(rows):
    """Native row-by-row date scan."""
    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row):
            if not val or val == "None": continue
            
            # 1. Direct cell value check
            d = parse_date_value(val)
            if d: return d
            
            # 2. Keyword check
            s_val = str(val).upper()
            if any(k in s_val for k in ["DATE", "AS AT", "PERIOD ENDING", "FOR THE"]):
                # Check right cell
                if c_idx + 1 < len(row):
                    d_r = parse_date_value(row[c_idx+1])
                    if d_r: return d_r
                # Check below cell (Peep next row)
                if r_idx + 1 < len(rows):
                    d_d = parse_date_value(rows[r_idx+1][c_idx])
                    if d_d: return d_d
    return None

def search_date_in_text(text):
    patterns = [
        r'(\d{2})[-/](\d{2})[-/](\d{4})', # 30-01-2026
        r'(\d{4})[-/](\d{2})[-/](\d{2})'  # 2026-01-30
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            try:
                v1, v2, v3 = int(m.group(1)), int(m.group(2)), int(m.group(3))
                if v1 > 1000: return datetime.date(v1, v2, v3) # YYYY-MM-DD
                return datetime.date(v3, v2, v1) # DD-MM-YYYY
            except: pass
    return None

def parse_date_value(val):
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.date() if hasattr(val, 'date') else val
    # Handle float dates from xlrd (Excel serial dates)
    if isinstance(val, float) and val > 40000:
        try: return datetime.date(1899, 12, 30) + datetime.timedelta(days=val)
        except: pass
    if val:
        return search_date_in_text(str(val))
    return None

def is_actually_bsd4(filepath):
    return analyze_return(filepath)['type'] in ['BSD4', 'BOTH']

def is_valid_bsd_return(filepath):
    return analyze_return(filepath)['type'] in ['BSD2_3', 'BOTH']

def is_valid_bsd4_return(filepath):
    return analyze_return(filepath)['type'] in ['BSD4', 'BOTH']


