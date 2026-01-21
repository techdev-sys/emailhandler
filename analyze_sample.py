
import openpyxl
import os

file_path = r"c:\Users\chinogs\Music\RBZ_Auto_Bot\ACL_2026-01-14.xlsx"

print(f"Analyzing {file_path}...")

if not os.path.exists(file_path):
    print("File not found!")
else:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print("\nSheets found:")
        for sheet in wb.sheetnames:
            print(f"- {sheet}")
            ws = wb[sheet]
            print(f"  First 5 rows of {sheet}:")
            for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=5):
                row_vals = [str(cell.value).strip() if cell.value else "None" for cell in row]
                print(f"    {row_vals}")
    except Exception as e:
        print(f"Error: {e}")
